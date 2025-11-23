import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os
import time

class Product:
    def __init__(self, name=None, category=None, price=None, offer_price=None, description=None, product_url=None):
        self.name = name or ""
        self.category = category or ""
        self.price = price or 0
        self.offer_price = offer_price or 0
        self.description = description or ""
        self.product_url = product_url or ""

    def to_list(self):
        return [self.name, self.category, self.price, self.offer_price, self.description, self.product_url]

# Initialize Excel file
def initialize_excel(filename="DCC_bazar_products.xlsx"):
    if os.path.exists(filename):
        # File exists, load it
        wb = load_workbook(filename)
        ws = wb.active
        print(f"✓ Continuing with existing file: {filename}")
        print(f"✓ Current rows in Excel: {ws.max_row}")
    else:
        # Create new file with headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = ["Name", "Category", "Price", "Offer Price", "Description", "Product URL"]
        ws.append(headers)
        wb.save(filename)
        print(f"✓ Created new file: {filename}")
    
    return wb, ws, filename

# Save single product to Excel
def save_single_product(product, wb, ws, filename):
    try:
        ws.append(product.to_list())
        wb.save(filename)
        print(f"✓ Saved to Excel: {product.name}")
        return True
    except Exception as e:
        print(f"✗ Error saving to Excel: {str(e)}")
        return False

# List of categories to scrape
categories = ['baby-food','baby-bath-skin','toys-and-games','baby-diaper','feeding-items','baby-grooming','walkers-and-strollers']
base_url = 'https://www.dccbazar.com/'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36'
}

# Initialize Excel file
wb, ws, excel_filename = initialize_excel()
all_products = []

for category in categories:
    print(f"\n=== Scraping category: {category} ===")
    
    category_url = f"{base_url}{category}.html?limit=all"
    
    try:
        r = requests.get(category_url, headers=headers)
        soup = BeautifulSoup(r.content, 'lxml')

        productlist = soup.find_all('div', class_='category-products-grid')
        productURLs = []

        for item in productlist:
            url = item.find('div',class_='images-container').find('div',class_='actions-no hover-box').find('a')
            if url and 'href' in url.attrs:
                productURLs.append(url['href'])

        print(f"Found {len(productURLs)} products in {category} category")

        for i, link in enumerate(productURLs):
            try:
                print(f"Processing product {i+1}/{len(productURLs)} in {category}")
                
                r = requests.get(link, headers=headers)
                soup = BeautifulSoup(r.content, 'lxml')

                # Get product name with error handling
                try:
                    p_name = soup.find('div', class_='product-name').find('h1').text.strip()
                except (AttributeError, IndexError):
                    p_name = ""

                # Get product description with error handling
                try:
                    desc_container = soup.find('div', class_='box-collateral box-description') \
                                    or soup.find('div', class_='box-collateral box-description active')

                    std_block = desc_container.find('div', class_='std') if desc_container else None

                    if std_block:
                        # Extract all text inside <div class="std"> including <p>, <br>, <li>, etc.
                        p_desc = std_block.get_text("\n", strip=True)
                    else:
                        p_desc = ""

                except:
                    p_desc = ""

                # Get prices with error handling
                try:
                    p_price = soup.find('span',class_='price').text.strip()
                except (AttributeError, IndexError):
                    p_price = ""

                # Create product with category and URL
                product = Product(
                    name=p_name,
                    description=p_desc,
                    price=p_price,
                    offer_price=p_price,
                    category=category.replace('-', ' ').title(),  # Convert 'skin-care' to 'Skin Care'
                    product_url=link  # Add product URL
                )

                # Save immediately to Excel
                success = save_single_product(product, wb, ws, excel_filename)
                if success:
                    all_products.append(product)
                    print(f"✓ Successfully scraped and saved: {p_name}")
                else:
                    print(f"✗ Failed to save: {p_name}")

                # Add a small delay to be respectful to the server
                # time.sleep(1)

            except Exception as e:
                print(f"✗ Error processing {link}: {str(e)}")
                continue

    except Exception as e:
        print(f"✗ Error accessing category {category}: {str(e)}")
        continue

print(f"\n=== Scraping Complete ===")
print(f"Total products scraped across all categories: {len(all_products)}")

# Print summary by category
from collections import defaultdict
category_summary = defaultdict(int)
for product in all_products:
    category_summary[product.category] += 1

print("\n=== Category Summary ===")
for category, count in category_summary.items():
    print(f"{category}: {count} products")

print(f"\nFinal data saved to: {excel_filename}")