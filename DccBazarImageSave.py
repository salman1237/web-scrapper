import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os
import time
from urllib.parse import urljoin, urlparse
import re

class Product:
    def __init__(self, name=None, category=None, price=None, offer_price=None, description=None, product_url=None, image_path=None):
        self.name = name or ""
        self.category = category or ""
        self.price = price or 0
        self.offer_price = offer_price or 0
        self.description = description or ""
        self.product_url = product_url or ""
        self.image_path = image_path or ""

    def to_list(self):
        return [self.name, self.category, self.price, self.offer_price, self.description, self.product_url, self.image_path]

# Initialize Excel file with image column
def initialize_excel(filename="DCC_bazar_products.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        print(f"✓ Continuing with existing file: {filename}")
        print(f"✓ Current rows in Excel: {ws.max_row}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = ["Name", "Category", "Price", "Offer Price", "Description", "Product URL", "Image Path"]
        ws.append(headers)
        wb.save(filename)
        print(f"✓ Created new file: {filename}")
    
    return wb, ws, filename

# Create images directory
def create_images_directory():
    images_dir = "product_images"
    if not os.path.exists(images_dir):
        os.makedirs(images_dir)
        print(f"✓ Created directory: {images_dir}")
    return images_dir

# Download image and return local path
def download_product_image(img_url, product_name, category, images_dir):
    try:
        if not img_url:
            return ""
        
        # Clean product name for filename
        clean_name = re.sub(r'[<>:"/\\|?*]', '', product_name)[:50]  # Remove invalid chars and limit length
        clean_name = clean_name.strip()
        
        if not clean_name:
            clean_name = "unknown_product"
        
        # Get file extension from URL
        parsed_url = urlparse(img_url)
        file_ext = os.path.splitext(parsed_url.path)[1]
        if not file_ext:
            file_ext = ".jpg"  # Default extension
        
        # Create filename
        filename = f"{clean_name}_{int(time.time())}{file_ext}"
        filepath = os.path.join(images_dir, category, filename)
        
        # Create category subdirectory if it doesn't exist
        category_dir = os.path.join(images_dir, category)
        if not os.path.exists(category_dir):
            os.makedirs(category_dir)
        
        # Download image
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36'
        }
        
        # Fix URL if needed (remove spaces and fix typos)
        img_url = img_url.replace(' ', '')  # Remove spaces from URL
        img_url = img_url.replace('..', '.')  # Fix double dots
        
        response = requests.get(img_url, headers=headers, timeout=30)
        response.raise_for_status()
        
        # Save image
        with open(filepath, 'wb') as f:
            f.write(response.content)
        
        print(f"✓ Downloaded image: {filename}")
        return filepath
        
    except Exception as e:
        print(f"✗ Error downloading image: {str(e)}")
        return ""

# Extract image URL from product page - UPDATED FOR YOUR HTML STRUCTURE
def extract_image_url(soup):
    try:
        # Method 1: Look for the main product image with class 'gallery-image visible'
        main_image = soup.find('img', class_='gallery-image visible')
        if main_image and main_image.get('src'):
            img_url = main_image['src']
            return fix_image_url(img_url)
        
        # Method 2: Look for image with id 'image-main'
        image_main = soup.find('img', id='image-main')
        if image_main and image_main.get('src'):
            img_url = image_main['src']
            return fix_image_url(img_url)
        
        # Method 3: Look for any image in product-image-gallery div
        gallery = soup.find('div', class_='product-image-gallery')
        if gallery:
            images = gallery.find_all('img')
            for img in images:
                if img.get('src'):
                    img_url = img['src']
                    return fix_image_url(img_url)
        
        # Method 4: Look for product image in product-img-content div
        img_content = soup.find('div', class_='product-img-content')
        if img_content:
            images = img_content.find_all('img')
            for img in images:
                if img.get('src'):
                    img_url = img['src']
                    return fix_image_url(img_url)
        
        # Method 5: Look for meta og:image
        meta_image = soup.find('meta', property='og:image')
        if meta_image and meta_image.get('content'):
            img_url = meta_image['content']
            return fix_image_url(img_url)
            
        return ""
    except Exception as e:
        print(f"✗ Error extracting image URL: {str(e)}")
        return ""

# Fix common URL issues in the image URLs
def fix_image_url(img_url):
    if not img_url:
        return ""
    
    # Fix the URL based on the issues I see in your HTML
    img_url = img_url.strip()
    
    # Fix double dots and spaces
    img_url = img_url.replace(' ', '')
    img_url = img_url.replace('..', '.')
    
    # Fix common typos in the domain
    img_url = img_url.replace('ww.dccbazar..con', 'www.dccbazar.com')
    img_url = img_url.replace('ww.dccbazar.con', 'www.dccbazar.com')
    img_url = img_url.replace('mm.dccbazar.con', 'www.dccbazar.com')
    
    # Ensure proper protocol
    if img_url.startswith('//'):
        img_url = 'https:' + img_url
    elif img_url.startswith('/'):
        img_url = 'https://www.dccbazar.com' + img_url
    
    return img_url

# Clean price - remove symbols and keep only numbers
def clean_price(price_text):
    if not price_text:
        return "0"
    
    # Remove all non-numeric characters except decimal point
    cleaned = re.sub(r'[^\d.]', '', str(price_text))
    
    # If empty after cleaning, return "0"
    if not cleaned:
        return "0"
    
    return cleaned

# Save single product to Excel
def save_single_product(product, wb, ws, filename):
    try:
        ws.append(product.to_list())
        wb.save(filename)
        print(f"✓ Saved to Excel: {product.name}")
        return True
    except Exception as e:
        print(f"✗ Error saving to Excel: {str(e)}")
        return False

# Main scraping code
categories = ['baby-food','baby-bath-skin','toys-and-games','baby-diaper','feeding-items','baby-grooming','walkers-and-strollers']
base_url = 'https://www.dccbazar.com/'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36'
}

# Initialize
wb, ws, excel_filename = initialize_excel()
images_dir = create_images_directory()
all_products = []

for category in categories:
    print(f"\n=== Scraping category: {category} ===")
    
    category_url = f"{base_url}{category}.html?limit=all"
    
    try:
        r = requests.get(category_url, headers=headers)
        soup = BeautifulSoup(r.content, 'lxml')

        productlist = soup.find_all('div', class_='category-products-grid')
        productURLs = []

        for item in productlist:
            url = item.find('div',class_='images-container').find('div',class_='actions-no hover-box').find('a')
            if url and 'href' in url.attrs:
                productURLs.append(url['href'])

        print(f"Found {len(productURLs)} products in {category} category")

        for i, link in enumerate(productURLs):
            try:
                print(f"Processing product {i+1}/{len(productURLs)} in {category}")
                
                r = requests.get(link, headers=headers)
                soup = BeautifulSoup(r.content, 'lxml')

                # Get product name
                try:
                    p_name = soup.find('div', class_='product-name').find('h1').text.strip()
                except (AttributeError, IndexError):
                    p_name = ""

                # Get product description
                try:
                    desc_container = soup.find('div', class_='box-collateral box-description') \
                                    or soup.find('div', class_='box-collateral box-description active')

                    std_block = desc_container.find('div', class_='std') if desc_container else None

                    if std_block:
                        p_desc = std_block.get_text("\n", strip=True)
                    else:
                        p_desc = ""
                except:
                    p_desc = ""

                # Get prices and clean them
                try:
                    p_price = soup.find('span',class_='price').text.strip()
                    p_price = clean_price(p_price)
                except (AttributeError, IndexError):
                    p_price = "0"

                # Extract and download image
                img_url = extract_image_url(soup)
                local_image_path = ""
                
                if img_url and p_name:
                    print(f"  Found image URL: {img_url}")
                    local_image_path = download_product_image(img_url, p_name, category, images_dir)
                else:
                    print("  No image found for this product")

                # Create product
                product = Product(
                    name=p_name,
                    description=p_desc,
                    price=p_price,
                    offer_price=p_price,
                    category=category.replace('-', ' ').title(),
                    product_url=link,
                    image_path=local_image_path
                )

                # Save to Excel
                success = save_single_product(product, wb, ws, excel_filename)
                if success:
                    all_products.append(product)
                    print(f"✓ Successfully scraped and saved: {p_name}")
                else:
                    print(f"✗ Failed to save: {p_name}")

                # Optional: Add delay to be respectful to server
                #time.sleep(1)  # Added delay to avoid overwhelming the server

            except Exception as e:
                print(f"✗ Error processing {link}: {str(e)}")
                continue

    except Exception as e:
        print(f"✗ Error accessing category {category}: {str(e)}")
        continue

print(f"\n=== Scraping Complete ===")
print(f"Total products scraped: {len(all_products)}")

# Print summary
from collections import defaultdict
category_summary = defaultdict(int)
for product in all_products:
    category_summary[product.category] += 1

print("\n=== Category Summary ===")
for category, count in category_summary.items():
    print(f"{category}: {count} products")

print(f"\nFinal data saved to: {excel_filename}")
print(f"Images saved to: {images_dir}")